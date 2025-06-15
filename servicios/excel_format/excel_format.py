from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

class ExcelFormat():

    HOJAS = ['Clientes', 'Ventas', 'Tickets', 'Inventario', 'Agrupado']

    def __init__(self, path: str):
        self.__path = path
        self.archivo = self.__check_path()

    def format(self):

        wb = load_workbook(self.archivo)

        for i in self.HOJAS:
            ws = wb[i]

            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill("solid", fgColor="4F81BD")
            center_align = Alignment(horizontal='center')

            for col_idx, cell in enumerate(ws[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(12, len(cell.value) + 4)


            for cell in ws['C'][1:]:  
                cell.number_format = '$#,##0.00'

        wb.save(self.archivo)

    def __check_path(self):
        try:
            ruta_relativa = os.path.join(os.path.dirname(__file__), f'../../{self.__path}')
            excel_path = os.path.abspath(ruta_relativa)
        except Exception as ex:
            raise Exception(f'No pude leer el path por esta razon: {ex}')
        
        return excel_path